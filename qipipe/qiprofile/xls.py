import re
from openpyxl import load_workbook
from contextlib import contextmanager
from datetime import datetime
from bunch import bunchify
import inflection
from qiutil.logging import logger
from .parsers import parse_trailing_number


class XLSError(Exception):
    pass


class Reader(object):
    """Reads a clinical Excel workbook file object."""

    def __init__(self, filename, sheet, **parsers):
        """
        :param filename: the attr:`filename`
        :param parsers: the {field: parser} options
        """
        self.filename = filename
        """The input XLS workbook file location."""

        self.parsers = parsers
        """The {attribute: value parser function} dictionary,"""
        
        # The workbook.
        wb = load_workbook(filename, read_only=True)
        # The worksheet.
        ws = wb[sheet]
        # The openpyxls row iterator.
        base_iter = ws.iter_rows()

        # The first row is the headings. After pulling off this row,
        # the base iterator resumes at the first data row.
        headings = next(base_iter)
        # The attributes in the same order as the headings.
        self.attributes = [self._attributize(cell.value)
                           for cell in headings]
        """The row attributes."""
        
        # The row bunch iterator. Empty rows are skipped.
        self._rows = (bunchify(self._parse_row(row)) for row in base_iter
                      if any(cell.value != None for cell in row))

    def _parse_row(self, row):
        """Extracts and parses the row values."""
        attr_val_dict = {}
        for i, cell in enumerate(row):
            attr = self.attributes[i]
            parser = self.parsers.get(attr, None)
            attr_val_dict[attr] = self._parse_cell(cell, parser)
        
        return attr_val_dict

    def _parse_cell(self, cell, parser=None):
        """
        Extracts and parses the given XLS cell.
        
        :param value: the XLS cell
        :param parser: the cell value parser
        """
        if cell.value and parser:
            return parser(cell.value)

    def __iter__(self):
        """
        Converts each input XLS field into a lower-case, underscore
        attribute and parses each input XLS row field value.

        :yield: the XLS row {attribute: parsed value} dictionary
        """
        for row in self._rows:
            yield row

    def filter(self, subject, session=None):
        """
        :param subject: the XNAT subject name
        :param session: the XNAT session name (required only if the
            session field is in the file)
        :yield: the :meth:`next` row
        """
        if not subject:
            raise XLSError("The subject search target is missing")
        # The subject attribute is required.
        sbj_attr = next((attr for attr in self.attributes
                         if attr == 'subject' or attr == 'patient'),
                        None)
        if not sbj_attr:
            raise XLSError("The Excel workbook file does not have a Subject"
                           " or Patient column header in the Excel workbook"
                           " file %s" % self.filename)
        # If there is a session search value, then the session attribute
        # is required.
        if session:
            sess_attr = next((attr for attr in self.attributes
                              if attr == 'session' or attr == 'visit'),
                             None)
            if not sess_attr:
                raise XLSError("Excel workbook file does not have a Session"
                               " or Visit column header in the Excel workbook"
                               " file %s" % self.filename)

        # The target subject number.
        tgt_sbj_num = self._extract_trailing_number(subject)
        # The target session number.
        tgt_sess_num = self._extract_trailing_number(session) if session else None

        # Apply the filter to each row.
        for row in self:
            # Match on the subject number.
            row_sbj = row.pop(sbj_attr)
            # Skip empty rows.
            if not row_sbj:
                continue
            # Extract the subject number.
            row_sbj_nbr = self._extract_trailing_number(row_sbj)
            # Check the row subject number against the target.
            if row_sbj_nbr != tgt_sbj_num:
                # No match: skip this row.
                continue
            # The row subject number matches the target. If there is a
            # session target, then match on the session number as well.
            if tgt_sess_num:
                row_sess = row.pop(sess_attr)
                # If we got this far, then there should be a session number.
                if not row_sess:
                    raise XLSError("The row for subject number %d is missing"
                                   " a session number." % tgt_sess_num)
                row_sess_nbr = self._extract_trailing_number(row_sess)
                if row_sess_nbr != tgt_sess_num:
                    # No match; skip this row.
                    continue
            # We have a winner.
            yield row

    def _extract_trailing_number(self, value):
        if isinstance(value, int):
            return value
        elif isinstance(value, str):
            return parse_trailing_number(value)
        else:
            raise XLSError("Cannot extract a trailing number from the value"
                           " %s" % value)

    def _attributize(self, s):
        """
        Converts the given XLS field name to an underscore attribute.

        :param s: the string to convert, or None if no cell value
        :return: the underscore attribute name, or None if the input
            string is None
        """
        if s:
            return inflection.underscore(re.sub(r'\W+', '_', s))
